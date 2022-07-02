#!/usr/bin/env python 
# -*- coding:utf-8 -*-
import time

import openpyxl

from common.path_util import pathutil
from common.request_util import requestsutil


class excelutil():

    # 创建excel表名
    timestamp = time.strftime("%Y%m%d", time.localtime())
    excelname = '%s测试报告' % timestamp
    suffix='.xlsx'
    savepath=pathutil().get_general_path()+'\\excel_report\\'+ excelname+suffix

    # 读取定义的表格用例,并返回一个用例列表(用例格式为字典格式)
    """
    filename为文件名称,sheetname为表格中sheet表
    """
    def read_excel(self, filename, sheetname):
        wb = openpyxl.load_workbook(filename)  # 加载excel文件
        sheet = wb[sheetname]  # 获取sheet/获取表单
        max_raw = sheet.max_row  # 获取选中的表单中最大的行数
        case_list = []  # 获取测试用例
        for i in range(2, max_raw + 1):
            dict1 = dict(
                id=sheet.cell(row=i, column=1).value,  # 获取excel的id
                url=sheet.cell(row=i, column=2).value,
                name=sheet.cell(row=i, column=3).value,
                headers=sheet.cell(row=i, column=4).value,
                method=sheet.cell(row=i, column=5).value,
                params=sheet.cell(row=i, column=6).value,
                excepts=sheet.cell(row=i, column=7).value,

            )  # 将每一行接口数据写成字典格式
            case_list.append(dict1)  # 每循环一次就把每一行的接口内容放在case_list列表中
        return case_list  # 返回一个用例列表

    # 写入断言,测试结果,并保存
    """
     filename为文件名称,sheetname为表格中sheet表
    row代表某行,column某一列,values所写入的值
    """
    def write_excel(self, filename, sheetname, row, column, values):
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheetname]
        case_result = sheet.cell(row=row, column=column).value = values
        wb.save(filename)


    #此方法专为输出excel表格使用
    def creat_excel(self, file_name,row_data):
        """
        :param file_name: 表格名称
        :param row_data:  用例的结果数据列表
        :return:
        """
        #实例化一个表格对象
        wb=openpyxl.Workbook()
        #创建表格的名字,并切换到第一个sheet
        ws=wb.active
        ws.title='test_report'
        suffix = '.xlsx'


        #添加首行数据
        first_row=['用例名称','请求方法','请求参数','实际结果','断言结果']
        ws.append(first_row)

        # 添加数据
        try:

            for i in row_data:
                ws.append(i)
            wb.save(pathutil().get_general_path()+'\\excel_report\\'+ file_name+suffix)
        except:
            print('表格写入异常,请检查!')


    #加载表格的所有数据,可以用于查看表格的测试报告
    def load_exceldata(self,file_name,sheet_name):
        wb = openpyxl.load_workbook(file_name)  # 加载excel文件
        sheet = wb[sheet_name]  # 获取sheet/获取表单
        max_raw = sheet.max_row  # 获取选中的表单中最大的行数
        max_col = sheet.max_column #获取选中的表单中最大的列数
        data_list=[]
        for i in range(1,max_raw+1):
            for j in range(1,max_col+1):
                value=sheet.cell(i,j).value
                data_list.append(value)
        return data_list


    #以下create_Yaml_excel和write_data用于yaml用例的测试报告
    def create_Yaml_excel(self):
        # 实例化一个表格对象
        wb = openpyxl.Workbook()
        # 创建表格的名字,并切换到第一个sheet
        ws = wb.active
        #创建sheet的名字
        ws.title = 'test-report'
        #定义后缀类型
        suffix='.xlsx'

        #空表首行
        first_row = ['用例名称', '请求方法', '请求参数', '实际结果', '断言结果']
        ws.append(first_row)
        wb.save(excelutil().savepath)

    def write_data(self,data):
        wb = openpyxl.load_workbook(excelutil().savepath)
        sheet = wb['test-report']
        ws = wb.active
        try:
            ws.append(data)
            wb.save(excelutil().savepath)
        except:
            raise FileNotFoundError('表格数据写入失败,请检查数据格式!!!')

# if __name__ == '__main__':
#     cases = excelutil().read_excel('D:/PyCharm2019/APIframe/case_data/cases/Cases.xlsx', '1.登录')
#     for unit_case in cases:
#         case_id = unit_case.get("case_id")
#         case_url = unit_case.get("url")
#         case_name = unit_case.get("name")
#         case_method = unit_case.get("method")
#         case_headers = (unit_case.get("headers"))
#         case_data = (unit_case.get("case_data"))
#         print(unit_case)
        # req = requestsutil().send_request(case_method, url=case_url, headers=case_headers, case_data=case_data)
        # print(req)
    # excelutil().create_Yaml_excel()
    # data=[[1,2,3,4,5],[7,8,9,10,11],["",'15717823601','reponse :{"method": "get", "url": "http://poetry.apiopen.top/getTime", "body": "None"}','null','这里是接口的返回值:  True,  断言成功!!!'],{'user':"1571754405"}]
    # for i in data:
    #     excelutil().write_data(i)